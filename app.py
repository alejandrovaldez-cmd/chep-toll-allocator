"""
CHEP Toll Allocator — Streamlit Web App
========================================
Browser-based interface for the toll allocation tool. Users upload the four
vendor files (and optionally the Master Asset Document), choose run options,
and download the daily allocation Excel.

Run locally:    streamlit run app.py
Deploy:         Streamlit Community Cloud (free) or any container host
"""

import io
import calendar
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CHEP Toll Allocator",
    page_icon="🚛",
    layout="wide",
)


# ─── REFERENCE DATA (same logic as the Python script versions) ────────────────
CHEP_MARKETS = {
    "Albany, GA", "Atlanta, GA", "Baltimore, MD", "Cudahy, WI", "DePere, WI",
    "Las Vegas, NV", "Louisville, KY", "Maumelle, AR", "Nashville, TN",
    "Richmond, VA", "Roanoke, VA", "Reno, NV", "Tolleson, AZ", "Barrington, NJ",
    "Gresham, OR", "Hammond, LA", "Lakewood, WA", "Leetsdale, PA",
    "Los Angeles, CA", "Mechanicsburg, PA", "Olive Branch, MS", "Rochester, NY",
    "Suffolk, VA", "Tomah, WI", "Tunkhannock, PA", "Fresno, CA", "Denver, CO",
    "Tulsa, OK", "Fayetteville, AR", "Springfield, MO", "Harlingen, TX",
    "Waterford, NY", "Norwood, MA",
}

REGION_TO_CHEP_MARKET = {
    "Albany, GA": "Albany, GA",
    "Atlanta, GA": "Atlanta, GA",
    "Atlanta, GA - Dry Van": "Atlanta, GA",
    "Baltimore, MD": "Baltimore, MD",
    "Barrington, NJ": "Barrington, NJ",
    "Cudahy, WI": "Cudahy, WI",
    "De Pere, WI": "DePere, WI",
    "DePere, WI": "DePere, WI",
    "Denver, CO": "Denver, CO",
    "Aurora, CO": "Denver, CO",
    "Fayetteville, AR": "Fayetteville, AR",
    "Fresno, CA": "Fresno, CA",
    "Gresham, OR": "Gresham, OR",
    "Hammond, LA": "Hammond, LA",
    "Harlingen, TX": "Harlingen, TX",
    "Lakewood, WA": "Lakewood, WA",
    "Las Vegas, NV": "Las Vegas, NV",
    "Leetsdale, PA": "Leetsdale, PA",
    "Los Angeles, CA": "Los Angeles, CA",
    "Southern CA": "Los Angeles, CA",
    "Louisville, KY": "Louisville, KY",
    "Maumelle, AR": "Maumelle, AR",
    "Mechanicsburg, PA": "Mechanicsburg, PA",
    "Nashville, TN": "Nashville, TN",
    "Norwood, MA": "Norwood, MA",
    "Olive Branch, MS": "Olive Branch, MS",
    "Reno, NV": "Reno, NV",
    "Richmond, VA": "Richmond, VA",
    "Roanoke, VA": "Roanoke, VA",
    "Rochester, NY": "Rochester, NY",
    "Springfield, MO": "Springfield, MO",
    "Suffolk, VA": "Suffolk, VA",
    "Tolleson, AZ": "Tolleson, AZ",
    "Tomah, WI": "Tomah, WI",
    "Tulsa, OK": "Tulsa, OK",
    "Tunkhannock, PA": "Tunkhannock, PA",
    "Waterford, NY": "Waterford, NY",
}

ROW_OVERRIDES = {
    ("Bestpass", "ALMZ1065DV", "Virginia (Richmond, Roanoke, Suffolk)"): "Richmond, VA",
    ("Bestpass", "ALMZ8371DV", "Virginia Dry Van (VA)"):                 "Mechanicsburg, PA",
    ("XTRA",     "W92836",     "Virginia Dry Van (VA)"):                 "Mechanicsburg, PA",
    ("Bestpass", "ALMZ1378DV", "-"):                                     "Suffolk, VA",
    ("Bestpass", "ALMZ1381DV", "-"):                                     "Roanoke, VA",
}

SCAC_TO_CHEP_MARKET = {
    "AT": "Atlanta, GA",   "AY": "Albany, GA",       "AZ": "Tolleson, AZ",
    "BM": "Mechanicsburg, PA", "BT": "Barrington, NJ", "CD": "Cudahy, WI",
    "DN": "Denver, CO",    "DP": "DePere, WI",       "FN": "Fresno, CA",
    "FY": "Fayetteville, AR", "GR": "Gresham, OR",   "HG": "Harlingen, TX",
    "HM": "Hammond, LA",   "KY": "Louisville, KY",   "LA": "Los Angeles, CA",
    "LD": "Leetsdale, PA", "LK": "Lakewood, WA",     "LV": "Las Vegas, NV",
    "MB": "Mechanicsburg, PA", "ML": "Maumelle, AR", "NR": "Norwood, MA",
    "NS": "Nashville, TN", "OB": "Olive Branch, MS", "PA": "Tunkhannock, PA",
    "RC": "Rochester, NY", "RI": "Richmond, VA",     "RN": "Reno, NV",
    "RO": "Roanoke, VA",   "SK": "Suffolk, VA",      "SP": "Springfield, MO",
    "TM": "Tomah, WI",     "TU": "Tulsa, OK",        "WF": "Waterford, NY",
    "SR": "Waterford, NY",
}

UNUSABLE_REGIONS = {"Unknown", "Reserved/Transit", "-", "", "nan"}


# ─── VENDOR LOADERS ───────────────────────────────────────────────────────────
def _read_excel_robust(file_bytes, **kwargs):
    """Try multiple Excel engines for resilience against malformed files.

    Bestpass portal exports sometimes produce .xlsx files that fail with the
    default openpyxl engine (e.g., 'There is no item named xl/sharedStrings.xml
    in the archive'). The calamine engine is more lenient and handles these
    cases. We fall through several engines before giving up.
    """
    engines_to_try = ["openpyxl", "calamine", "xlrd"]
    last_error = None
    for engine in engines_to_try:
        try:
            if hasattr(file_bytes, "seek"):
                file_bytes.seek(0)
            return pd.read_excel(file_bytes, engine=engine, **kwargs)
        except ImportError:
            # Engine not installed — skip silently
            continue
        except Exception as e:
            last_error = e
            continue
    # Re-raise the most recent real error so the user sees something useful
    raise last_error if last_error else RuntimeError("No working Excel engine available")


def load_premier(file_bytes):
    df = _read_excel_robust(file_bytes, sheet_name="Toll Lines")
    df["Vendor"] = "Premier"
    df["UnitID"] = df["Equip ID"].astype(str)
    df["Amount"] = pd.to_numeric(df["Line Total"], errors="coerce").fillna(0)
    df["InvoiceDate"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    df["CostCenter"] = None
    return df

def _read_sheet_flexible(file_bytes, candidates, header=0):
    """Try a list of possible sheet names and return the first that loads.
    Useful when the same vendor's file has different sheet names depending on
    whether it's a raw export vs an accounting-pre-mapped file."""
    last_error = None
    # Reset file pointer between attempts (uploaded files are seekable)
    for sheet_name in candidates:
        try:
            if hasattr(file_bytes, "seek"):
                file_bytes.seek(0)
            return _read_excel_robust(file_bytes, sheet_name=sheet_name, header=header)
        except Exception as e:
            last_error = e
            continue
    raise ValueError(
        f"None of these sheet names were found: {candidates}. "
        f"Last error: {last_error}"
    )


def load_star(file_bytes):
    # Try accounting's pre-mapped sheet first, then raw export's "Sheet1"
    df = _read_sheet_flexible(file_bytes, ["Invoice Lines", "Sheet1"])

    # Filter to toll lines (Manual Charges only)
    if "Charge Type" in df.columns:
        df = df[df["Charge Type"] == "MAN - Manual Charge"].copy()

    df["Vendor"] = "Star"
    df["UnitID"] = df["Unit Number"].astype(str)
    # Force numeric — Star raw exports sometimes store amounts as text strings
    df["Amount"] = pd.to_numeric(df["Invoice Line Total"], errors="coerce").fillna(0)
    # Date parsing: try the known format first, fall back to flexible parsing
    df["InvoiceDate"] = pd.to_datetime(
        df["Invoice Date"], format="%m-%d-%Y", errors="coerce"
    )
    if df["InvoiceDate"].isna().all():
        df["InvoiceDate"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    df["CostCenter"] = None
    return df

def load_xtra(file_bytes):
    df = _read_excel_robust(file_bytes, sheet_name="Invoice detail", header=2)
    df = df[df["Description"] == "Toll Fee"].copy()
    df["Vendor"] = "XTRA"
    df["UnitID"] = df["Unit #"].astype(str)
    df["VIN_lookup"] = df["VIN"].astype(str)
    df["Amount"] = pd.to_numeric(df["Line Total"], errors="coerce").fillna(0)
    df["InvoiceDate"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    df["CostCenter"] = None
    return df

def load_bestpass(file_bytes):
    df = _read_excel_robust(file_bytes, sheet_name="Toll Activity")
    df["Vendor"] = "Bestpass"
    df["UnitID"] = df["Unit"].astype(str)
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["InvoiceDate"] = pd.to_datetime(df["Post Date"], errors="coerce")
    df["CostCenter"] = df["Cost Center"]
    return df


VENDOR_LOADERS = {
    "Premier": load_premier,
    "Star": load_star,
    "XTRA": load_xtra,
    "Bestpass": load_bestpass,
}


# ─── VENDOR AUTO-DETECTION ────────────────────────────────────────────────────
# Filename keywords (case-insensitive) — these MUST be vendor-specific. Avoid
# anything generic like "monthly", "report", "invoice" — those collide with
# multiple vendors' export naming conventions.
FILENAME_HINTS = {
    "Premier":  ["premier", "ptlz"],
    "Star":     ["star_", "star leasing"],   # avoid bare "star" — too short, easy false-positives
    "XTRA":     ["xtra"],
    "Bestpass": ["bestpass", "fleetworthy"],
}

# Sheet-name fingerprints — strong signals that uniquely identify a vendor's file
SHEET_FINGERPRINTS = {
    "Premier":  ["Premier Data >", "Toll Lines"],
    "Star":     ["Star Data >"],          # accounting's pre-mapped Star file has this tab
    "XTRA":     ["Invoice detail"],
    "Bestpass": ["Toll Activity"],
}

# Column-header fingerprints — for files where filename and sheet name aren't
# helpful (e.g., raw exports with generic names like "Sheet1"). Each entry is
# a set of columns that, taken together, uniquely identify a vendor's data.
COLUMN_FINGERPRINTS = [
    # Premier: Equip ID + Toll columns are unique to Premier's billing data
    ("Premier", {"Equip ID", "Toll"}),
    ("Premier", {"Equip ID", "Admin Fee"}),
    ("Premier", {"Equip ID", "Toll Description"}),
    # Star: Charge Type + Unit Number + Invoice Line Total — Star-specific naming
    ("Star",    {"Charge Type", "Unit Number", "Invoice Line Total"}),
    ("Star",    {"Inv Source", "Charge Type", "Unit Number"}),
    # Bestpass: Cost Center + Transaction Desc + Agency
    ("Bestpass", {"Cost Center", "Transaction Desc"}),
    ("Bestpass", {"Cost Center", "Agency", "Plaza"}),
]


def detect_vendor(file_obj):
    """Identify which vendor a file belongs to.

    Strategy (in order of confidence):
      1. Sheet-name fingerprint — strongest signal (vendor-specific sheets)
      2. Column-header fingerprint — for raw exports w/ generic sheet names
      3. Filename keyword — fallback, only for clearly vendor-named files

    Returns (vendor_name, detection_method) or (None, error_message).
    """
    name = getattr(file_obj, "name", "").lower()

    # 1. Sheet-name fingerprint (most reliable)
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        # Try multiple engines for robustness — Bestpass portal exports can be
        # malformed in ways openpyxl rejects but calamine handles.
        xl = None
        for engine in ("openpyxl", "calamine", "xlrd"):
            try:
                if hasattr(file_obj, "seek"):
                    file_obj.seek(0)
                xl = pd.ExcelFile(file_obj, engine=engine)
                break
            except (ImportError, Exception):
                continue
        if xl is None:
            return (None, "could not open file with any Excel engine")

        sheets = set(xl.sheet_names)
        for vendor, fingerprints in SHEET_FINGERPRINTS.items():
            if any(fp in sheets for fp in fingerprints):
                return (vendor, "sheet name")

        # 2. Column-header fingerprint — peek at first sheet
        if xl.sheet_names:
            try:
                # Try header on row 0 first, then row 2 (XTRA-style)
                for hdr_row in (0, 2):
                    try:
                        df_peek = _read_excel_robust(
                            file_obj, sheet_name=xl.sheet_names[0],
                            header=hdr_row, nrows=2
                        )
                        cols = {str(c).strip() for c in df_peek.columns
                                if isinstance(c, str)}
                        for vendor, required_cols in COLUMN_FINGERPRINTS:
                            if required_cols.issubset(cols):
                                return (vendor, "column headers")
                    except Exception:
                        continue
            except Exception:
                pass
    except Exception as e:
        # Couldn't open file at all — fall through to filename check
        pass

    # 3. Filename hint — last resort (least reliable, since users rename files)
    for vendor, hints in FILENAME_HINTS.items():
        if any(hint in name for hint in hints):
            return (vendor, "filename")

    return (None, "could not identify vendor")


# ─── MAD LOOKUP BUILDER ───────────────────────────────────────────────────────
def build_mad_lookups(mad_file_bytes):
    mar = _read_excel_robust(mad_file_bytes, sheet_name="Master Asset Record")
    mar["_priority"] = mar["Status"].map(
        {"Active": 0, "Active Rental": 0,
         "Pending Return / Displacement Market ": 1, "Onboarding": 1, "Standby": 1}
    ).fillna(2)
    mar = mar.sort_values("_priority")

    def build(df, id_col):
        d = df.dropna(subset=[id_col]).copy()
        d["_k"] = d[id_col].astype(str).str.strip()
        d = d.drop_duplicates(subset=["_k"], keep="first")
        return d.set_index("_k")[
            ["Customer", "Current Region", "Originating Market"]
        ].to_dict(orient="index")

    return {
        "internal": build(mar, "Internal ID"),
        "leasing":  build(mar, "Leasing ID"),
        "vin":      build(mar, "Vehicle Identification Number"),
    }


def lookup_asset_mad(unit_id, vin, primary_lookup_name, mad):
    primary = mad[primary_lookup_name]
    unit_norm = str(unit_id).strip() if pd.notna(unit_id) else ""
    if unit_norm and unit_norm in primary:
        info = primary[unit_norm]
        return (info["Customer"], info["Current Region"], info["Originating Market"], "matched")
    if vin and pd.notna(vin):
        vin_norm = str(vin).strip()
        if vin_norm in mad["vin"]:
            info = mad["vin"][vin_norm]
            return (info["Customer"], info["Current Region"], info["Originating Market"], "vin_fallback")
    return (None, None, None, "no_match")


# ─── REGION → MARKET MAPPING ──────────────────────────────────────────────────
def parse_scac(value):
    if not value or pd.isna(value):
        return None
    s = str(value).strip().upper()
    if len(s) == 4 and s[:2] in ("VR", "VT") and s[2:] in SCAC_TO_CHEP_MARKET:
        return s[2:]
    return None


def map_to_market(vendor, unit_id, region, origin=None, cost_center=None):
    """Three-tier mapping with fallbacks."""
    region_str = str(region).strip() if pd.notna(region) else ""
    origin_str = str(origin).strip() if pd.notna(origin) else ""

    # Row override
    key = (vendor, str(unit_id).strip(), region_str)
    if key in ROW_OVERRIDES:
        return (ROW_OVERRIDES[key], "OK (row override)")

    # Direct region lookup
    if region_str and region_str not in UNUSABLE_REGIONS and not region_str.startswith("#"):
        if region_str in REGION_TO_CHEP_MARKET:
            return (REGION_TO_CHEP_MARKET[region_str], "OK (region)")

    # Origin fallback
    if origin_str and origin_str not in UNUSABLE_REGIONS and not origin_str.startswith("#"):
        origin_key = (vendor, str(unit_id).strip(), origin_str)
        if origin_key in ROW_OVERRIDES:
            return (ROW_OVERRIDES[origin_key], "OK (override on origin)")
        if origin_str in REGION_TO_CHEP_MARKET:
            return (REGION_TO_CHEP_MARKET[origin_str], f"OK (origin: '{origin_str}')")

    # Bestpass Cost Center fallback
    if vendor == "Bestpass" and cost_center and pd.notna(cost_center):
        cc_str = str(cost_center).strip()
        if cc_str in REGION_TO_CHEP_MARKET:
            return (REGION_TO_CHEP_MARKET[cc_str], f"OK (Cost Center: '{cc_str}')")
        scac = parse_scac(cc_str)
        if scac:
            return (SCAC_TO_CHEP_MARKET[scac], f"OK (SCAC: '{cc_str}')")

    if region_str in UNUSABLE_REGIONS or not region_str:
        if origin_str in UNUSABLE_REGIONS or not origin_str:
            return ("REVIEW", "Both Region and Origin missing/unusable")
        return ("REVIEW", f"Origin '{origin_str}' not in mapping")
    return ("REVIEW", f"Region '{region_str}' not in mapping")


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
LABEL_FONT  = Font(bold=True, name="Arial", size=11)
BODY_FONT   = Font(name="Arial", size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def auto_size(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel_output(allocations, exceptions, all_data, no_match_data,
                      cost_label, bill_label, n_days, day_dates, mode_label):
    wb = Workbook()

    # Tab 1: Daily Allocation
    ws1 = wb.active
    ws1.title = "Daily Allocation"
    ws1["A1"] = f"CHEP Toll Allocation — {cost_label} → billed across {bill_label}"
    ws1["A1"].font = Font(bold=True, size=14, name="Arial", color="305496")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_days + 2)
    ws1["A2"] = f"Mode: {mode_label}  |  Billable days: {n_days}"
    ws1["A2"].font = Font(italic=True, size=10, name="Arial", color="595959")

    headers = ["CHEP Market", "Monthly Total"] + [d.strftime("%m/%d") for d in day_dates]
    hdr_row = 4
    for c, h in enumerate(headers, 1):
        ws1.cell(row=hdr_row, column=c, value=h)
    style_header(ws1, hdr_row, len(headers))

    sorted_alloc = allocations.sort_values("CHEP Market").reset_index(drop=True)
    for i, row in sorted_alloc.iterrows():
        r = hdr_row + 1 + i
        ws1.cell(row=r, column=1, value=row["CHEP Market"]).font = BODY_FONT
        c2 = ws1.cell(row=r, column=2, value=float(row["Monthly Total"]))
        c2.font = BODY_FONT; c2.number_format = '"$"#,##0.00'
        for d in range(n_days):
            wc = ws1.cell(row=r, column=3 + d, value=f"=$B{r}/{n_days}")
            wc.font = BODY_FONT; wc.number_format = '"$"#,##0.0000'
        for c in range(1, len(headers) + 1):
            ws1.cell(row=r, column=c).border = BORDER

    total_row = hdr_row + 1 + len(sorted_alloc)
    ws1.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
    ws1.cell(row=total_row, column=2, value=f"=SUM(B{hdr_row+1}:B{total_row-1})")
    for c in range(3, n_days + 3):
        col = get_column_letter(c)
        ws1.cell(row=total_row, column=c, value=f"=SUM({col}{hdr_row+1}:{col}{total_row-1})")
    for c in range(1, len(headers) + 1):
        cell = ws1.cell(row=total_row, column=c)
        cell.fill = TOTAL_FILL; cell.font = LABEL_FONT; cell.border = BORDER
        if c >= 2:
            cell.number_format = '"$"#,##0.00' if c == 2 else '"$"#,##0.0000'
    auto_size(ws1, [24, 14] + [11] * n_days)
    ws1.freeze_panes = "C5"

    # Tab 2: Monthly Summary
    ws2 = wb.create_sheet("Monthly Summary")
    ws2["A1"] = "CHEP Toll Costs by Market × Vendor"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color="305496")
    chep_data = all_data[all_data["CHEP Market"] != "REVIEW"]
    pivot = chep_data.pivot_table(index="CHEP Market", columns="Vendor",
                                   values="Amount", aggfunc="sum", fill_value=0)
    vendors = sorted(pivot.columns.tolist()) if len(pivot) else []
    headers = ["CHEP Market"] + vendors + ["Total"]
    hdr_row = 3
    for c, h in enumerate(headers, 1):
        ws2.cell(row=hdr_row, column=c, value=h)
    style_header(ws2, hdr_row, len(headers))

    psort = pivot.sort_index()
    for i, (mkt, row) in enumerate(psort.iterrows()):
        r = hdr_row + 1 + i
        ws2.cell(row=r, column=1, value=mkt).font = BODY_FONT
        for j, v in enumerate(vendors):
            cell = ws2.cell(row=r, column=2 + j, value=float(row[v]))
            cell.font = BODY_FONT; cell.number_format = '"$"#,##0.00'
        first = get_column_letter(2); last = get_column_letter(1 + len(vendors))
        tot = ws2.cell(row=r, column=len(headers), value=f"=SUM({first}{r}:{last}{r})")
        tot.number_format = '"$"#,##0.00'; tot.font = LABEL_FONT
        for c in range(1, len(headers) + 1):
            ws2.cell(row=r, column=c).border = BORDER
    if len(psort):
        total_row = hdr_row + 1 + len(psort)
        ws2.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
        for c in range(2, len(headers) + 1):
            col = get_column_letter(c)
            cell = ws2.cell(row=total_row, column=c,
                            value=f"=SUM({col}{hdr_row+1}:{col}{total_row-1})")
            cell.fill = TOTAL_FILL; cell.font = LABEL_FONT
            cell.number_format = '"$"#,##0.00'; cell.border = BORDER
        ws2.cell(row=total_row, column=1).fill = TOTAL_FILL
        ws2.cell(row=total_row, column=1).border = BORDER
    auto_size(ws2, [24] + [14] * len(vendors) + [14])
    ws2.freeze_panes = "B4"

    # Tab 3: Exceptions
    ws3 = wb.create_sheet("Exceptions")
    ws3["A1"] = "Exceptions — Manual Review Required"
    ws3["A1"].font = Font(bold=True, size=14, name="Arial", color="C00000")

    # Combine no-match (MAD-only) and review exceptions
    exc_rows = []
    if no_match_data is not None and len(no_match_data):
        for _, r in no_match_data.iterrows():
            exc_rows.append({
                "Vendor": r["Vendor"], "UnitID": r["UnitID"],
                "Customer": "(no match)", "Region": "(no match)",
                "Amount": r["Amount"], "Reason": "Unit not in MAD"
            })
    for _, r in exceptions.iterrows():
        exc_rows.append({
            "Vendor": r["Vendor"], "UnitID": r["UnitID"],
            "Customer": str(r.get("Customer", r.get("MAD_Customer", ""))),
            "Region": str(r.get("Region", r.get("MAD_Region", ""))),
            "Amount": r["Amount"], "Reason": r["Reason"]
        })

    if exc_rows:
        headers = ["Vendor", "Unit / VIN", "Customer", "Region", "Amount", "Reason"]
        hdr_row = 3
        for c, h in enumerate(headers, 1):
            ws3.cell(row=hdr_row, column=c, value=h)
        style_header(ws3, hdr_row, len(headers))
        for i, row in enumerate(exc_rows):
            r = hdr_row + 1 + i
            ws3.cell(row=r, column=1, value=row["Vendor"]).font = BODY_FONT
            ws3.cell(row=r, column=2, value=row["UnitID"]).font = BODY_FONT
            ws3.cell(row=r, column=3, value=row["Customer"]).font = BODY_FONT
            ws3.cell(row=r, column=4, value=row["Region"]).font = BODY_FONT
            amt = ws3.cell(row=r, column=5, value=float(row["Amount"]))
            amt.font = BODY_FONT; amt.number_format = '"$"#,##0.00'
            ws3.cell(row=r, column=6, value=row["Reason"]).font = BODY_FONT
            for c in range(1, 7):
                ws3.cell(row=r, column=c).border = BORDER
                ws3.cell(row=r, column=c).fill = ALERT_FILL
        auto_size(ws3, [12, 18, 16, 28, 14, 50])
    else:
        ws3["A3"] = "No exceptions — all CHEP toll lines mapped successfully ✓"
        ws3["A3"].font = Font(italic=True, color="548235", size=12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─── PROCESSING ───────────────────────────────────────────────────────────────
def process(vendor_files, mad_file, mode, cost_year, cost_month, bill_year, bill_month,
            filter_by_invoice_date):
    """Run the full pipeline, return (allocations_df, exceptions_df, all_data,
    no_match_df, summary_dict)."""
    frames = []
    summary = {
        "vendors_loaded": [],
        "deferred": {},  # vendor -> count, amount
    }

    # Load MAD lookups if MAD-driven mode
    mad = None
    if mode == "MAD-driven" and mad_file is not None:
        mad = build_mad_lookups(mad_file)

    for vendor, file in vendor_files.items():
        if file is None:
            continue
        try:
            df = VENDOR_LOADERS[vendor](file)
        except Exception as e:
            st.error(f"Failed to load {vendor}: {e}")
            continue

        # Customer + Region: either from file's columns or from MAD
        if mode == "Pre-mapped" or mad is None:
            # Use the Customer / Region columns from the source file
            # If the columns don't exist (raw export), this vendor's data
            # can't be used in Pre-mapped mode — flag and skip cleanly.
            has_customer = "Customer" in df.columns
            has_region   = "Region" in df.columns

            if not has_customer or not has_region:
                summary.setdefault("missing_premapped_cols", []).append({
                    "vendor": vendor,
                    "missing": [c for c in ["Customer", "Region"]
                                if c not in df.columns],
                })
                # Skip this vendor's data in Pre-mapped mode
                continue

            df["Customer_resolved"] = df["Customer"]
            df["Region_resolved"]   = df["Region"]
            df["Origin_resolved"]   = None
            df["match_type"] = "pre_mapped"
        else:
            # MAD-driven: look up each row
            primary = "leasing" if vendor in ("Premier", "Star", "XTRA") else "internal"
            results = df.apply(
                lambda r: lookup_asset_mad(
                    r["UnitID"],
                    r.get("VIN_lookup") if "VIN_lookup" in df.columns else None,
                    primary, mad
                ),
                axis=1, result_type="expand"
            )
            df["Customer_resolved"] = results[0]
            df["Region_resolved"]   = results[1]
            df["Origin_resolved"]   = results[2]
            df["match_type"]        = results[3]

        # Filter to CHEP only
        df = df[df["Customer_resolved"].astype(str).str.strip() == "CHEP"].copy()

        # Filter by invoice date if requested
        if filter_by_invoice_date:
            in_month = (df["InvoiceDate"].dt.year == cost_year) & \
                       (df["InvoiceDate"].dt.month == cost_month)
            deferred = df[~in_month]
            summary["deferred"][vendor] = (len(deferred), deferred["Amount"].sum())
            df = df[in_month]

        df["VendorFile"] = vendor
        summary["vendors_loaded"].append({
            "vendor": vendor,
            "rows": len(df),
            "amount": df["Amount"].sum() if len(df) else 0,
        })
        frames.append(df)

    if not frames:
        return None, None, None, None, summary

    all_data = pd.concat(frames, ignore_index=True)

    # Map regions to markets — use result_type="expand" to handle
    # small/single-row dataframes correctly
    if len(all_data) == 0:
        all_data["CHEP Market"] = []
        all_data["Reason"] = []
    else:
        mapping_results = all_data.apply(
            lambda r: pd.Series(map_to_market(
                r["Vendor"], r["UnitID"],
                r["Region_resolved"], r["Origin_resolved"], r["CostCenter"]
            )),
            axis=1, result_type="expand"
        )
        # Defensively rename columns in case apply returned them by integer index
        if len(mapping_results.columns) == 2:
            mapping_results.columns = ["CHEP Market", "Reason"]
        all_data["CHEP Market"] = mapping_results["CHEP Market"]
        all_data["Reason"] = mapping_results["Reason"]

    # Track no-match rows separately
    no_match_data = pd.DataFrame()
    if mode == "MAD-driven":
        no_match_data = all_data[all_data["match_type"] == "no_match"].copy()
        all_data = all_data[all_data["match_type"] != "no_match"]

    # Exceptions are rows that mapped to REVIEW
    exceptions = all_data[all_data["CHEP Market"] == "REVIEW"][
        ["Vendor", "UnitID", "Customer_resolved", "Region_resolved", "Amount", "Reason"]
    ].rename(columns={"Customer_resolved": "Customer", "Region_resolved": "Region"})

    # Allocations are CHEP rows that mapped to a real market
    allocatable = all_data[all_data["CHEP Market"] != "REVIEW"]
    allocations = allocatable.groupby("CHEP Market", as_index=False)["Amount"] \
                              .sum() \
                              .rename(columns={"Amount": "Monthly Total"})

    return allocations, exceptions, all_data, no_match_data, summary


# ─── UI ───────────────────────────────────────────────────────────────────────
st.title("🚛 CHEP Toll Allocator")
st.caption("Process monthly trailer-leasing toll files and produce CHEP daily bill-back allocations.")

with st.expander("ℹ️ How this works", expanded=False):
    st.markdown("""
**Inputs:**
- Four vendor files (Premier, Star, XTRA, Bestpass) — Excel
- Master Asset Document (only required for MAD-driven mode)

**What it does:**
1. Filters each vendor's data to toll-only line items
2. (MAD mode) Looks up each unit in the Master Asset Document to derive Customer + Region
3. Filters to CHEP customers only
4. (Optional) Filters by invoice date — keeps only rows invoiced in the cost month
5. Maps each row's region to one of the 33 CHEP Dedicated Lane markets, with three-tier fallback (Region → Origin → Bestpass Cost Center / SCAC)
6. Aggregates totals by market and divides across the days of the billing month
7. Outputs a daily allocation Excel with reconciliation info and any unmapped exceptions

**Modes:**
- **MAD-driven** — the tool does its own Customer/Region lookups using the Master Asset Document. No accounting pre-mapping needed.
- **Pre-mapped** — uses whatever Customer/Region columns are already in the vendor files (e.g., from accounting's pre-processing).
""")

# ─── Sidebar: settings ─────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")

    mode = st.radio(
        "Processing mode",
        options=["MAD-driven", "Pre-mapped"],
        index=0,
        help="MAD-driven does its own lookups. Pre-mapped uses Customer/Region from the vendor files."
    )

    st.divider()
    st.subheader("Cost Month")
    st.caption("Month whose vendor invoices are being processed")
    today = datetime.now()
    default_cost_month = today.month - 1 if today.month > 1 else 12
    default_cost_year = today.year if today.month > 1 else today.year - 1
    cost_year = st.number_input("Cost year", min_value=2020, max_value=2050,
                                value=default_cost_year, step=1)
    cost_month = st.selectbox("Cost month",
                              options=list(range(1, 13)),
                              index=default_cost_month - 1,
                              format_func=lambda m: calendar.month_name[m])

    st.divider()
    st.subheader("Billing Month")
    st.caption("Month CHEP is being billed in (output spans this month)")
    bill_year = st.number_input("Billing year", min_value=2020, max_value=2050,
                                value=today.year, step=1)
    bill_month = st.selectbox("Billing month",
                              options=list(range(1, 13)),
                              index=today.month - 1,
                              format_func=lambda m: calendar.month_name[m])

    st.divider()
    filter_by_invoice = st.checkbox(
        "Filter by Invoice Date",
        value=True,
        help="Match accounting's methodology: only include rows where the vendor invoice date falls in the cost month. Rows invoiced in subsequent months are deferred to the next billing cycle."
    )

# ─── Main: file uploads ───────────────────────────────────────────────────────
st.subheader("📁 Upload Vendor Files")
st.caption(
    "Drop in any of the four vendor files (Premier, Star, XTRA, Bestpass) — "
    "the app will figure out which is which automatically. You can drag multiple files at once."
)
uploaded_vendor_files = st.file_uploader(
    "Vendor files",
    type=["xlsx"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key="vendor_files",
)

# Auto-detect each uploaded file's vendor
detected = {}      # vendor name → file object
detection_rows = []  # rows for the user-visible detection table

if uploaded_vendor_files:
    for f in uploaded_vendor_files:
        vendor, method = detect_vendor(f)
        # Reset the file's read position after detection (loader will read it again)
        if hasattr(f, "seek"):
            f.seek(0)

        if vendor and vendor not in detected:
            detected[vendor] = f
            detection_rows.append({
                "File": f.name,
                "Detected as": vendor,
                "How": f"via {method}",
                "Status": "✅",
            })
        elif vendor and vendor in detected:
            # Duplicate — same vendor uploaded twice
            detection_rows.append({
                "File": f.name,
                "Detected as": vendor,
                "How": f"via {method}",
                "Status": f"⚠️ Duplicate — using earlier {vendor} file instead",
            })
        else:
            detection_rows.append({
                "File": f.name,
                "Detected as": "—",
                "How": method,  # error message
                "Status": "❌ Unrecognized — file will be skipped",
            })

    st.dataframe(pd.DataFrame(detection_rows), hide_index=True, use_container_width=True)

    # Show what's still missing
    missing = [v for v in ["Premier", "Star", "XTRA", "Bestpass"] if v not in detected]
    if missing:
        st.caption(f"Still missing: {', '.join(missing)}")

# Bind the detected files to the variable names the rest of the app expects
f_premier  = detected.get("Premier")
f_star     = detected.get("Star")
f_xtra     = detected.get("XTRA")
f_bestpass = detected.get("Bestpass")

st.subheader("📋 Master Asset Document")
if mode == "MAD-driven":
    st.caption("Required for MAD-driven mode. Open the Master Asset Document → download the 'Master Asset Record' tab as Excel → upload below.")
    st.link_button(
        "🔗 Open Master Asset Document",
        "https://docs.google.com/spreadsheets/d/1KdWMMGn-2tJghHhX8LBg_-Mve5omcD_rOqJp1C-b3Cw/edit?gid=288332535#gid=288332535"
    )
else:
    st.caption("Optional — only used in MAD-driven mode.")
f_mad = st.file_uploader("Vorto CPG Master Asset Document", type=["xlsx"], key="mad")

st.divider()

# ─── Run button ───────────────────────────────────────────────────────────────
ready = (f_premier or f_star or f_xtra or f_bestpass)
if mode == "MAD-driven":
    ready = ready and f_mad is not None

# Determine which vendors are present and which are missing
vendors_present = [v for v, f in [
    ("Premier", f_premier), ("Star", f_star),
    ("XTRA", f_xtra), ("Bestpass", f_bestpass)
] if f is not None]
vendors_missing = [v for v in ["Premier", "Star", "XTRA", "Bestpass"]
                   if v not in vendors_present]
is_partial_run = len(vendors_present) > 0 and len(vendors_missing) > 0

if not ready:
    if mode == "MAD-driven":
        st.info("Upload at least one vendor file and the Master Asset Document to get started.")
    else:
        st.info("Upload at least one vendor file to get started.")
else:
    # Show partial-run warning before they click Run
    if is_partial_run:
        st.warning(
            f"⚠️ **Partial run** — only {len(vendors_present)} of 4 vendors uploaded.\n\n"
            f"**Included:** {', '.join(vendors_present)}\n\n"
            f"**Missing:** {', '.join(vendors_missing)}\n\n"
            f"This is fine for spot-checking a single vendor's CHEP totals. "
            f"**If you're producing a bill-back run, make sure you don't double-bill** "
            f"when the missing vendors are added later."
        )

    if st.button("▶️ Run Allocation", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            vendor_files = {
                "Premier": f_premier, "Star": f_star,
                "XTRA": f_xtra, "Bestpass": f_bestpass,
            }

            allocations, exceptions, all_data, no_match, summary = process(
                vendor_files, f_mad, mode,
                cost_year, cost_month, bill_year, bill_month,
                filter_by_invoice
            )

            if allocations is None or len(allocations) == 0:
                # Check if Pre-mapped mode failed because raw files were used
                if mode == "Pre-mapped" and summary.get("missing_premapped_cols"):
                    skipped = summary["missing_premapped_cols"]
                    skipped_names = ", ".join(s["vendor"] for s in skipped)
                    st.error(
                        f"**Pre-mapped mode can't process these vendors: {skipped_names}**\n\n"
                        f"Their files don't have the `Customer` and `Region` columns that "
                        f"Pre-mapped mode requires (which accounting normally adds during pre-processing).\n\n"
                        f"**Solution:** Switch to **MAD-driven mode** in the sidebar — it works "
                        f"with raw vendor files by deriving Customer/Region from the Master Asset Document."
                    )
                else:
                    st.error("No CHEP toll data found. Check your files and settings.")
            else:
                # Summary metrics
                if is_partial_run:
                    st.success(f"✅ Processed successfully — partial run ({len(vendors_present)}/4 vendors)")
                    st.warning(
                        f"⚠️ **This is a partial run.** "
                        f"Included: {', '.join(vendors_present)}. "
                        f"Missing: {', '.join(vendors_missing)}. "
                        f"Output Excel is labeled as partial. **Don't double-bill** when the missing vendors are added later."
                    )
                else:
                    st.success(f"✅ Processed successfully")

                # Warn about any vendors skipped in Pre-mapped mode
                if mode == "Pre-mapped" and summary.get("missing_premapped_cols"):
                    skipped = summary["missing_premapped_cols"]
                    skipped_names = ", ".join(s["vendor"] for s in skipped)
                    st.warning(
                        f"⚠️ **Skipped in Pre-mapped mode: {skipped_names}** — these files "
                        f"don't have Customer/Region columns. Switch to MAD-driven mode to include them."
                    )

                # Force numeric in case sparse data returned mixed/object dtype
                total = float(pd.to_numeric(allocations["Monthly Total"],
                                            errors="coerce").fillna(0).sum())
                n_days = calendar.monthrange(bill_year, bill_month)[1]
                cost_label = f"{calendar.month_name[cost_month]} {cost_year}"
                bill_label = f"{calendar.month_name[bill_month]} {bill_year}"

                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total CHEP Tolls", f"${total:,.2f}")
                col2.metric("Markets Allocated", len(allocations))
                col3.metric("Per-Day Average",
                            f"${(total/n_days) if n_days else 0:,.2f}")
                col4.metric("Exceptions", len(exceptions))

                # Vendor breakdown
                st.subheader(f"📊 Per-Vendor Breakdown ({cost_label})")
                if summary["vendors_loaded"]:
                    vd = pd.DataFrame(summary["vendors_loaded"])
                    vd["amount"] = pd.to_numeric(vd["amount"], errors="coerce").fillna(0)
                    vd["amount"] = vd["amount"].apply(lambda x: f"${float(x):,.2f}")
                    vd.columns = ["Vendor", "CHEP Toll Lines", "Total Amount"]
                    st.dataframe(vd, hide_index=True, use_container_width=True)

                # Show deferred amounts if filter was on
                if filter_by_invoice and any(v[0] > 0 for v in summary["deferred"].values()):
                    with st.expander("⏭️ Deferred to next billing cycle", expanded=False):
                        st.caption("Rows where the vendor invoice date is outside the cost month")
                        for v, (n, amt) in summary["deferred"].items():
                            if n > 0:
                                st.write(f"**{v}:** {n} rows / ${float(amt):,.2f}")

                # Allocation preview
                st.subheader(f"💰 CHEP Market Allocation → billed across {bill_label} ({n_days} days)")
                disp = allocations.sort_values("CHEP Market").copy()
                disp["Monthly Total"] = pd.to_numeric(disp["Monthly Total"], errors="coerce").fillna(0)
                disp["Per Day"] = disp["Monthly Total"] / n_days
                disp["Monthly Total"] = disp["Monthly Total"].apply(lambda x: f"${float(x):,.2f}")
                disp["Per Day"] = disp["Per Day"].apply(lambda x: f"${float(x):,.2f}")
                st.dataframe(disp, hide_index=True, use_container_width=True)

                # Exceptions
                if len(exceptions) > 0:
                    with st.expander(f"⚠️ Exceptions ({len(exceptions)} rows / ${exceptions['Amount'].sum():,.2f})", expanded=True):
                        st.caption("Rows that couldn't be mapped to a CHEP market — review and adjust manually.")
                        st.dataframe(exceptions, hide_index=True, use_container_width=True)

                if no_match is not None and len(no_match) > 0:
                    with st.expander(f"❓ Units not in MAD ({len(no_match)} rows / ${no_match['Amount'].sum():,.2f})"):
                        st.caption("Toll lines whose Unit#/VIN didn't match anything in the Master Asset Document. Usually driver-incurred or admin charges.")
                        st.dataframe(
                            no_match[["Vendor", "UnitID", "Amount"]],
                            hide_index=True, use_container_width=True
                        )

                # Build & offer Excel download
                day_dates = pd.date_range(
                    start=f"{bill_year}-{bill_month:02d}-01",
                    periods=n_days, freq="D"
                )
                # Append vendor list to mode label so it shows up in the Excel
                mode_with_vendors = mode
                if is_partial_run:
                    mode_with_vendors = (
                        f"{mode} — PARTIAL RUN "
                        f"({', '.join(vendors_present)} only; "
                        f"missing {', '.join(vendors_missing)})"
                    )
                excel_bytes = build_excel_output(
                    allocations, exceptions, all_data, no_match,
                    cost_label, bill_label, n_days, day_dates, mode_with_vendors
                )
                if is_partial_run:
                    fname = (f"CHEP_Toll_Allocation_PARTIAL_"
                             f"{'-'.join(vendors_present)}_"
                             f"{calendar.month_abbr[bill_month]}{bill_year}.xlsx")
                else:
                    fname = (f"CHEP_Toll_Allocation_"
                             f"{calendar.month_abbr[bill_month]}{bill_year}.xlsx")
                st.download_button(
                    "📥 Download Excel",
                    data=excel_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )

st.divider()
st.caption("Built for Vorto Finance Ops · Internal use only")
